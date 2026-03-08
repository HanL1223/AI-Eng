"""
sttm_loader.py — Smart STTM Workbook Loader
=============================================
Understands the structure of your Source-To-Target-Mapping workbook
and extracts clean, structured text that Claude can actually search.

This file also serves as a Python refresher — every technique is explained.

Usage:
    from sttm_loader import load_sttm_workbook
    documents = load_sttm_workbook("docs/STTM.xlsx")
"""

import os
import re

# ─────────────────────────────────────────────────────────────
# PYTHON REFRESHER: Type hints
# ─────────────────────────────────────────────────────────────
# The -> list[dict] means "this function returns a list of dictionaries"
# The param: str means "this parameter should be a string"
# These don't enforce anything — they're documentation for you and your IDE


def load_sttm_workbook(filepath: str) -> list[dict]:
    """
    Load an STTM Excel workbook and return structured documents.

    Each entity sheet (Fact_*, Dim_*, Bridge_*) becomes TWO documents:
      1. A "table summary" with the header metadata (grain, source, refresh, etc.)
      2. A "column mapping" with the column-level details

    Why two? Because when someone asks "what is the grain of Fact_Sales?",
    you want the retriever to find the summary — not a random column row.
    And when someone asks "where does StoreKey come from?", you want
    the column mapping — not the header metadata.

    Returns:
        list[dict]: Each dict has {"content": str, "source": str}
    """
    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: Import inside function
    # ─────────────────────────────────────────────────────────
    # Importing here means the rest of your code doesn't break
    # if openpyxl isn't installed — the error only happens when
    # you actually try to load an Excel file.
    try:
        import openpyxl
    except ImportError:
        print("  ❌ Install openpyxl: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    documents = []
    filename = os.path.basename(filepath)

    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: Iterating with enumerate()
    # ─────────────────────────────────────────────────────────
    # enumerate() gives you (index, value) pairs
    # The _ means "I don't need this variable" (common Python convention)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip tiny sheets (empty placeholders) and control sheets
        if ws.max_row <= 2:
            continue
        if sheet_name.lower() in ("control table", "review"):
            # But DO load the control table as a special document
            control_doc = _extract_control_table(ws, sheet_name, filename)
            if control_doc:
                documents.append(control_doc)
            continue

        # ─────────────────────────────────────────────────────
        # Extract this entity sheet
        # ─────────────────────────────────────────────────────
        summary_doc, columns_doc = _extract_entity_sheet(ws, sheet_name, filename)

        if summary_doc:
            documents.append(summary_doc)
            print(f"  📋 {sheet_name}: summary ({len(summary_doc['content'])} chars)")
        if columns_doc:
            documents.append(columns_doc)
            print(f"  📊 {sheet_name}: {columns_doc['column_count']} columns ({len(columns_doc['content'])} chars)")

    wb.close()
    return documents


def _extract_control_table(ws, sheet_name: str, filename: str) -> dict | None:
    """
    Extract the Control Table as a structured document.

    PYTHON REFRESHER: Functions starting with _ are "private by convention"
    — Python doesn't enforce it, but it signals "this is an internal helper,
    don't call it from outside this file."
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row (contains "Object Name" or "Object Number")
    header_idx = None
    for i, row in enumerate(rows):
        # ─────────────────────────────────────────────────────
        # PYTHON REFRESHER: any() with generator expression
        # ─────────────────────────────────────────────────────
        # any() returns True if ANY item in the iterable is True
        # This is equivalent to a for loop with a break, but more Pythonic:
        #   for cell in row:
        #       if cell and "object" in str(cell).lower():
        #           header_idx = i; break
        if any(cell and "object" in str(cell).lower() for cell in row):
            header_idx = i
            break

    if header_idx is None:
        return None

    headers = [_clean(str(c)) if c else f"col_{j}" for j, c in enumerate(rows[header_idx])]

    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: List comprehension with condition
    # ─────────────────────────────────────────────────────────
    # [expression FOR item IN iterable IF condition]
    # This builds a new list by transforming and filtering in one line
    lines = []
    for row in rows[header_idx + 1:]:
        # Skip empty rows
        if not any(cell for cell in row):
            continue

        # ─────────────────────────────────────────────────────
        # PYTHON REFRESHER: zip() pairs up two lists
        # ─────────────────────────────────────────────────────
        # zip(["a","b","c"], [1,2,3]) → [("a",1), ("b",2), ("c",3)]
        # We use it to pair header names with cell values
        parts = []
        for header, cell in zip(headers, row):
            if cell and str(cell).strip():
                val = _clean(str(cell))
                parts.append(f"{header}: {val}")

        if parts:
            # ─────────────────────────────────────────────────
            # PYTHON REFRESHER: str.join()
            # ─────────────────────────────────────────────────
            # " | ".join(["a", "b", "c"]) → "a | b | c"
            # The opposite of .split() — combines a list into a string
            lines.append(" | ".join(parts))

    content = f"STTM Control Table — List of all entities\n\n" + "\n".join(lines)
    return {"content": content, "source": f"{filename}__Control_Table"}


def _extract_entity_sheet(ws, sheet_name: str, filename: str) -> tuple:
    """
    Extract an entity sheet (Fact_*, Dim_*, Bridge_*) into two documents:
    1. Table summary (header metadata)
    2. Column mappings (column-level detail)

    PYTHON REFRESHER: -> tuple means this returns multiple values.
    In Python, you can return multiple values separated by commas:
        return value1, value2
    The caller unpacks them:
        summary, columns = _extract_entity_sheet(...)
    """
    rows = list(ws.iter_rows(values_only=True))

    # ─────────────────────────────────────────────────────────
    # PART 1: Extract header metadata (rows 0-12ish)
    # ─────────────────────────────────────────────────────────
    # Your STTM has a consistent structure:
    #   Row 0: Object Name | <value>
    #   Row 1: Description | <value>
    #   Row 3: Data Grain/Primary Key | <value>
    #   Row 4: Data Sources | <value>
    #   ...etc
    # We look for known labels in the first column

    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: Dictionaries
    # ─────────────────────────────────────────────────────────
    # dict() or {} creates a key-value store
    # metadata["grain"] = "Date + Store + Item"
    # metadata.get("grain", "unknown") returns "unknown" if key doesn't exist
    metadata = {}
    header_row_idx = None  # Where the column mapping headers start

    # These are the labels we expect in column A of the header section
    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: Dictionary for label mapping
    # ─────────────────────────────────────────────────────────
    # Maps what we find in the Excel → what we call it in our output
    label_map = {
        "object name": "table_name",
        "description": "description",
        "type": "table_type",
        "data grain": "grain",
        "data grain/primary key": "grain",
        "data sources": "data_sources",
        "legacy transformation": "transformation_logic",
        "load type": "load_type",
        "data classification": "data_classification",
        "contains pii": "contains_pii",
        "refresh interval": "refresh_interval",
        "specified refesh time": "refresh_time",   # note: typo is in original
        "specified refresh time": "refresh_time",
        "data steward": "data_steward",
        "no.of records": "record_count",
    }

    for i, row in enumerate(rows):
        first_cell = _clean(str(row[0])) if row[0] else ""
        first_lower = first_cell.lower()

        # Check if this row is a header label
        matched = False
        for label_prefix, key in label_map.items():
            # ─────────────────────────────────────────────────
            # PYTHON REFRESHER: str.startswith()
            # ─────────────────────────────────────────────────
            # "hello world".startswith("hello") → True
            # More reliable than == for matching, since labels
            # sometimes have extra text after them
            if first_lower.startswith(label_prefix):
                # The value is usually in column B (index 1)
                # But sometimes it spans multiple columns
                values = [_clean(str(c)) for c in row[1:] if c and _clean(str(c))]
                if values:
                    metadata[key] = " | ".join(values)
                matched = True
                break

        # Detect where column mapping headers start
        # (looks for "Column Name" or "Presentation" in a row)
        if any(c and "column name" in str(c).lower() for c in row):
            header_row_idx = i
            break
        if any(c and "presentation" in str(c).lower() for c in row):
            # The actual headers are the NEXT row
            if i + 1 < len(rows):
                header_row_idx = i + 1
            break

    # Build the summary document
    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: f-strings (formatted string literals)
    # ─────────────────────────────────────────────────────────
    # f"Hello {name}" inserts the variable 'name' into the string
    # f"{count:,}" formats with commas: 1340570946 → "1,340,570,946"
    table_name = metadata.get("table_name", sheet_name)
    summary_parts = [f"Table: {table_name}"]

    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: dict.items() for iteration
    # ─────────────────────────────────────────────────────────
    # .items() gives you (key, value) pairs to loop over
    display_labels = {
        "description": "Description",
        "table_type": "Type",
        "grain": "Grain / Primary Key",
        "data_sources": "Data Sources",
        "transformation_logic": "Transformation Logic",
        "load_type": "Load Type",
        "data_classification": "Data Classification",
        "contains_pii": "Contains PII",
        "refresh_interval": "Refresh Interval",
        "refresh_time": "Specified Refresh Time",
        "record_count": "Record Count",
        "data_steward": "Data Steward / Business Contact",
    }

    for key, label in display_labels.items():
        value = metadata.get(key)
        if value:
            summary_parts.append(f"{label}: {value}")

    summary_content = "\n".join(summary_parts)
    summary_doc = {
        "content": summary_content,
        "source": f"{filename}__{sheet_name}__summary",
    } if len(summary_parts) > 1 else None  # Don't create if only the table name

    # ─────────────────────────────────────────────────────────
    # PART 2: Extract column mappings
    # ─────────────────────────────────────────────────────────
    columns_doc = None
    if header_row_idx is not None and header_row_idx < len(rows) - 1:
        # Get the header row
        header_row = rows[header_row_idx]
        headers = []
        for j, cell in enumerate(header_row):
            if cell:
                headers.append((j, _clean(str(cell))))
            # Stop at reasonable number of columns
            if j > 20:
                break

        # ─────────────────────────────────────────────────────
        # PYTHON REFRESHER: List of dicts pattern
        # ─────────────────────────────────────────────────────
        # Very common in data engineering:
        #   columns = [{"name": "StoreKey", "type": "FK"}, ...]
        # Each dict represents a "row" with named fields
        column_lines = []

        for row in rows[header_row_idx + 1:]:
            # The first non-empty cell in the column-name position = the column name
            col_name = _clean(str(row[0])) if row[0] else ""
            if not col_name:
                continue  # Skip rows with no column name

            # Build a description of this column
            parts = [f"  {col_name}"]
            for col_idx, col_header in headers[1:]:  # Skip first (Column Name itself)
                if col_idx < len(row) and row[col_idx]:
                    val = _clean(str(row[col_idx]))
                    if val:
                        # ─────────────────────────────────────
                        # PYTHON REFRESHER: str.lower() for normalization
                        # ─────────────────────────────────────
                        # Normalize header names for cleaner output
                        h = col_header.lower()

                        if "data type" in h or "column data" in h:
                            parts.append(f"type={val}")
                        elif "definition" in h:
                            parts.append(f"desc=\"{val}\"")
                        elif "source system" in h:
                            parts.append(f"src_system={val}")
                        elif "source object" in h or "source table" in h:
                            parts.append(f"src_table={val}")
                        elif "source column" in h or "derivation" in h:
                            parts.append(f"src_col={val}")
                        elif "source db" in h:
                            parts.append(f"src_db={val}")
                        elif "source schema" in h:
                            parts.append(f"src_schema={val}")
                        elif "join" in h:
                            parts.append(f"join=\"{val[:100]}\"")
                        elif "etl" in h or "rule" in h:
                            parts.append(f"etl=\"{val[:100]}\"")
                        elif "type" in h and "surr" in h.lower():
                            parts.append(f"key_type={val}")
                        elif "scd" in h:
                            parts.append(f"scd={val}")
                        elif "friendly" in h:
                            parts.append(f"friendly_name={val}")
                        else:
                            parts.append(f"{col_header}={val}")

            column_lines.append(" | ".join(parts))

        if column_lines:
            columns_content = (
                f"Column Mapping for: {table_name}\n"
                f"Columns ({len(column_lines)}):\n"
                + "\n".join(column_lines)
            )
            columns_doc = {
                "content": columns_content,
                "source": f"{filename}__{sheet_name}__columns",
                "column_count": len(column_lines),
            }

    return summary_doc, columns_doc


def _clean(text: str) -> str:
    """
    Clean cell text: remove non-breaking spaces, extra whitespace, etc.

    PYTHON REFRESHER: Regular expressions (regex)
    ─────────────────────────────────────────────
    re.sub(pattern, replacement, string) finds all matches and replaces them.

    \\xa0 is a non-breaking space — common in Excel exports.
    \\s+  matches one or more whitespace characters (space, tab, newline).
    .strip() removes leading/trailing whitespace.
    """
    text = text.replace("\xa0", " ")       # non-breaking space → normal space
    text = text.replace("\\xa0", " ")      # sometimes it's escaped as literal string
    text = re.sub(r"\s+", " ", text)       # collapse multiple spaces/newlines
    text = text.strip()
    return text


# ─────────────────────────────────────────────────────────────
# TEST IT: Run this file directly to see the output
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    # ─────────────────────────────────────────────────────────
    # PYTHON REFRESHER: sys.argv (command line arguments)
    # ─────────────────────────────────────────────────────────
    # When you run: python sttm_loader.py myfile.xlsx
    # sys.argv = ["sttm_loader.py", "myfile.xlsx"]
    # sys.argv[0] is always the script name
    # sys.argv[1] is the first argument
    if len(sys.argv) < 2:
        print("Usage: python sttm_loader.py <path_to_sttm.xlsx>")
        print("Example: python sttm_loader.py docs/STTM.xlsx")
        sys.exit(1)

    filepath = sys.argv[1]
    print(f"\n📂 Loading STTM workbook: {filepath}")
    print("=" * 50)

    docs = load_sttm_workbook(filepath)

    print(f"\n✅ Extracted {len(docs)} documents")
    print()

    # Show first 3 documents as preview
    for i, doc in enumerate(docs[:6]):
        print(f"─── Document {i+1}: {doc['source']} ───")
        # ─────────────────────────────────────────────────────
        # PYTHON REFRESHER: String slicing
        # ─────────────────────────────────────────────────────
        # text[:300] gets the first 300 characters
        # text[10:20] gets characters 10 through 19
        # text[-5:] gets the last 5 characters
        print(doc["content"][:300])
        print("...\n")