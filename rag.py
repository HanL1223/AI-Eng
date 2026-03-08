"""
rag.py — RAG Chatbot (Week 1 Complete Version)
================================================
All improvements from the step-by-step guide already applied.
Drop this into your rag-chatbot/ folder and run: python rag.py

Changes from the starter version:
- Loads .txt, .md, AND .xlsx files
- Chunks include metadata (table_name, table_type)
- Auto-detects table names in your questions for filtered retrieval
- Domain-specific system prompt for Sigma Healthcare
"""

import os
import glob
import anthropic
import chromadb

# ─────────────────────────────────────────────
# CONFIG — tweak these as you learn
# ─────────────────────────────────────────────
DOCS_DIR = "docs"
CHROMA_DIR = "chroma_db"
COLLECTION_NAME = "my_docs"
CHUNK_SIZE = 500                     # TRY: 300, 500, 800, 1200 — see what works
CHUNK_OVERLAP = 50
TOP_K = 3                            # TRY: 1, 3, 5 — fewer = precise, more = broad
MODEL = "claude-sonnet-4-5-20250929"


# ─────────────────────────────────────────────
# STEP 1: Load documents (.txt, .md, .xlsx)
# ─────────────────────────────────────────────
def load_documents(docs_dir: str) -> list[dict]:
    """Read all .txt, .md, and .xlsx files from a directory."""
    documents = []

    # --- Text files ---
    for pattern in ["*.txt", "*.md"]:
        for filepath in glob.glob(os.path.join(docs_dir, pattern)):
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read()
            documents.append({
                "content": content,
                "source": os.path.basename(filepath),
            })
            print(f"  📄 Loaded: {os.path.basename(filepath)} ({len(content)} chars)")

    # --- Excel files ---
    for filepath in glob.glob(os.path.join(docs_dir, "*.xlsx")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            filename = os.path.basename(filepath)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_str = " | ".join(
                        [str(cell) if cell is not None else "" for cell in row]
                    )
                    if row_str.strip().replace("|", "").strip():
                        rows.append(row_str)

                if rows:
                    content = f"Sheet: {sheet_name}\n" + "\n".join(rows)
                    source_name = f"{filename}__{sheet_name}"
                    documents.append({
                        "content": content,
                        "source": source_name,
                    })
                    print(f"  📊 Loaded: {filename} → {sheet_name} ({len(content)} chars)")

            wb.close()
        except ImportError:
            print("  ⚠️  Install openpyxl to load Excel: pip install openpyxl")
            break
        except Exception as e:
            print(f"  ❌ Error loading {filepath}: {e}")

    return documents


# ─────────────────────────────────────────────
# STEP 2: Chunk documents with metadata tags
# ─────────────────────────────────────────────
def chunk_text(text: str, source: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[dict]:
    """
    Split text into overlapping chunks with metadata.

    Metadata extracted:
    - table_name: from the filename (DIM_PRODUCT.txt → DIM_PRODUCT)
    - table_type: guessed from prefix (DIM_ → dimension, FACT_ → fact, etc.)
    """
    # Extract table name from filename
    # Handle both "DIM_PRODUCT.txt" and "workbook__DIM_PRODUCT" (from Excel)
    base_name = os.path.splitext(source)[0]
    # If it's from Excel (has __), use the sheet name part
    if "__" in base_name:
        table_name = base_name.split("__")[-1]
    else:
        table_name = base_name

    # Guess table type from naming convention
    table_type = "unknown"
    upper = table_name.upper()
    if upper.startswith("DIM_"):
        table_type = "dimension"
    elif upper.startswith("FACT_"):
        table_type = "fact"
    elif upper.startswith("BRIDGE_"):
        table_type = "bridge"
    elif upper.startswith("STG_") or upper.startswith("BRZ_"):
        table_type = "staging"

    chunks = []
    start = 0

    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]

        if end < len(text):
            last_period = chunk.rfind(".")
            last_newline = chunk.rfind("\n")
            break_point = max(last_period, last_newline)
            if break_point > chunk_size * 0.3:
                chunk = chunk[: break_point + 1]
                end = start + break_point + 1

        chunks.append({
            "text": chunk.strip(),
            "source": source,
            "chunk_index": len(chunks),
            "table_name": table_name.upper(),
            "table_type": table_type,
        })

        start = end - overlap

    return chunks


# ─────────────────────────────────────────────
# STEP 3: Store chunks in ChromaDB
# ─────────────────────────────────────────────
def build_vector_store(chunks: list[dict]) -> chromadb.Collection:
    """Embed and store chunks with metadata in ChromaDB."""
    client = chromadb.PersistentClient(path=CHROMA_DIR)

    try:
        client.delete_collection(COLLECTION_NAME)
    except Exception:
        pass

    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"},
    )

    ids = []
    documents = []
    metadatas = []

    for i, chunk in enumerate(chunks):
        ids.append(f"chunk_{i}")
        documents.append(chunk["text"])
        metadatas.append({
            "source": chunk["source"],
            "chunk_index": chunk["chunk_index"],
            "table_name": chunk["table_name"],
            "table_type": chunk["table_type"],
        })

    collection.add(ids=ids, documents=documents, metadatas=metadatas)
    print(f"  💾 Stored {len(chunks)} chunks in ChromaDB")
    return collection


# ─────────────────────────────────────────────
# STEP 4: Retrieve with optional metadata filter
# ─────────────────────────────────────────────
def retrieve(collection: chromadb.Collection, query: str, top_k: int = TOP_K,
             table_name: str = None, table_type: str = None) -> list[dict]:
    """
    Find relevant chunks, optionally filtered by table name or type.
    If table_name is provided, only search chunks from that table.
    """
    # Build filter
    where_filter = None
    if table_name and table_type:
        where_filter = {"$and": [
            {"table_name": {"$eq": table_name}},
            {"table_type": {"$eq": table_type}},
        ]}
    elif table_name:
        where_filter = {"table_name": {"$eq": table_name}}
    elif table_type:
        where_filter = {"table_type": {"$eq": table_type}}

    query_params = {"query_texts": [query], "n_results": top_k}
    if where_filter:
        query_params["where"] = where_filter

    try:
        results = collection.query(**query_params)
    except Exception:
        # If filtered query fails (e.g. not enough results), fall back to unfiltered
        results = collection.query(query_texts=[query], n_results=top_k)

    retrieved = []
    for i in range(len(results["ids"][0])):
        retrieved.append({
            "text": results["documents"][0][i],
            "source": results["metadatas"][0][i]["source"],
            "table_name": results["metadatas"][0][i].get("table_name", ""),
            "table_type": results["metadatas"][0][i].get("table_type", ""),
            "distance": results["distances"][0][i] if results.get("distances") else None,
        })

    return retrieved


# ─────────────────────────────────────────────
# STEP 5: Auto-detect table name from question
# ─────────────────────────────────────────────
def extract_table_name(query: str, known_tables: list[str]) -> str | None:
    """
    If user asks about 'DIM_PRODUCT', detect it and filter retrieval.
    Matches against all table names loaded into the vector store.
    """
    query_upper = query.upper()
    # Sort by length descending so "FACT_SALES_DETAIL" matches before "FACT_SALES"
    for table in sorted(known_tables, key=len, reverse=True):
        if table in query_upper:
            return table
    return None


# ─────────────────────────────────────────────
# STEP 6: Ask Claude with retrieved context
# ─────────────────────────────────────────────
def ask_claude(query: str, context_chunks: list[dict]) -> str:
    """Send query + retrieved context to Claude."""
    context_parts = []
    for chunk in context_chunks:
        label = chunk["source"]
        if chunk.get("table_name"):
            label = f"{chunk['table_name']} ({chunk['source']})"
        context_parts.append(f"[Source: {label}]\n{chunk['text']}")
    context = "\n\n---\n\n".join(context_parts)

    client = anthropic.Anthropic()

    message = client.messages.create(
        model=MODEL,
        max_tokens=1024,
        system="""You are a data warehouse documentation assistant for Sigma Healthcare.

You answer questions about tables, columns, mappings, and data pipelines 
in Sigma's Snowflake data warehouse.

Context:
- The warehouse follows a Bronze → Platinum → Gold layer architecture
- Source systems include SAP (via CDS Views), and other operational systems
- Data flows: SAP CDS Views → Azure Data Factory → ADLS Gen2 → dbt → Snowflake
- Tables follow naming conventions: DIM_ (dimensions), FACT_ (facts), BRIDGE_ (bridges)

Rules:
- ONLY use information from the provided context documents to answer
- If the context doesn't contain the answer, say "I don't have that information in the loaded documents"
- Always cite which source file the information came from
- When describing a table, include: grain, key columns, source system, and refresh cadence if available
- When listing columns, format them clearly
- Be concise and direct — the user is a data engineer, not a business user""",
        messages=[
            {
                "role": "user",
                "content": f"""Context from documents:

{context}

---

Question: {query}""",
            }
        ],
    )

    return message.content[0].text


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("\n🔧 RAG Chatbot — Week 1 Complete Version")
    print("=" * 45)

    # Load API key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
        if os.path.exists(env_path):
            with open(env_path) as f:
                for line in f:
                    if line.strip() and not line.startswith("#"):
                        key, _, value = line.strip().partition("=")
                        os.environ[key] = value
        else:
            print("\n❌ No ANTHROPIC_API_KEY found!")
            print("   export ANTHROPIC_API_KEY='sk-ant-...'")
            return

    # Check docs
    if not os.path.exists(DOCS_DIR):
        print(f"\n❌ No '{DOCS_DIR}/' folder found! Create it and add files.")
        return

    # Step 1: Load
    print("\n📂 Loading documents...")
    documents = load_documents(DOCS_DIR)
    if not documents:
        print(f"  ⚠️  No files found in {DOCS_DIR}/")
        return

    # Step 2: Chunk
    print("\n✂️  Chunking documents...")
    all_chunks = []
    for doc in documents:
        chunks = chunk_text(doc["content"], doc["source"])
        all_chunks.extend(chunks)
    print(f"  📦 Created {len(all_chunks)} chunks from {len(documents)} documents")

    # Step 3: Store
    print("\n🗄️  Building vector store...")
    collection = build_vector_store(all_chunks)

    # Get known table names for auto-detection
    all_meta = collection.get()
    known_tables = list(set(
        m.get("table_name", "")
        for m in all_meta["metadatas"]
        if m.get("table_name") and m.get("table_name") != ""
    ))
    if known_tables:
        print(f"  📋 Known tables: {', '.join(sorted(known_tables))}")

    # Interactive loop
    print("\n✅ Ready! Ask questions about your documents.")
    print("   Commands: 'debug' toggle | 'tables' list | 'quit' exit\n")

    debug_mode = False

    while True:
        try:
            query = input("You: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n👋 Bye!")
            break

        if not query:
            continue
        if query.lower() == "quit":
            print("👋 Bye!")
            break
        if query.lower() == "debug":
            debug_mode = not debug_mode
            print(f"  🔍 Debug mode: {'ON' if debug_mode else 'OFF'}")
            continue
        if query.lower() == "tables":
            print(f"  📋 Tables loaded: {', '.join(sorted(known_tables))}")
            continue

        # Auto-detect table name from question
        detected_table = extract_table_name(query, known_tables)

        if debug_mode and detected_table:
            print(f"  🎯 Auto-detected table: {detected_table}")

        # Retrieve
        chunks = retrieve(collection, query, table_name=detected_table)

        if debug_mode:
            print(f"\n  📎 Retrieved {len(chunks)} chunks:")
            for i, c in enumerate(chunks):
                dist = f" (distance: {c['distance']:.4f})" if c["distance"] else ""
                tbl = f" [{c['table_name']}]" if c.get("table_name") else ""
                print(f"    [{i+1}]{tbl} {c['source']}{dist}")
                print(f"        {c['text'][:120]}...")
            print()

        # Generate answer
        try:
            answer = ask_claude(query, chunks)
            print(f"\nClaude: {answer}\n")
        except anthropic.AuthenticationError:
            print("\n❌ Invalid API key. Check your ANTHROPIC_API_KEY.")
            break
        except Exception as e:
            print(f"\n❌ Error: {e}")


if __name__ == "__main__":
    main()